"""
Perbandingan lengkap Excel Monitoring_Combine vs SQLite MonitoringSummary.
Fokus: Tahun 2026, Bulan Mei, satker bps1300-bps1304.
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook

# Setup Django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "intermilan_project.settings.development")
sys.path.insert(0, ".")
import django
django.setup()
from apps.core.models import MonitoringSummary

XLSX_PATH = Path(
    r"c:\Users\muall\Documents\INTERMILAN PROJECT anti gravity"
    r"\INTERMILAN PROJECT COPY\data_sources\Database awal"
    r"\drive-download-20260615T061835Z-3-001\INTERMILAN.xlsx"
)

TARGET_SATKERS = ["bps1300", "bps1301", "bps1302", "bps1303", "bps1304"]
TARGET_BULAN = "mei"
TARGET_TAHUN = 2026

# ================================================================
# BACA EXCEL
# ================================================================
print(f"Membuka: {XLSX_PATH}\n")
wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb["Monitoring_Combine"]

headers = None
excel_rows = {}  # key: satker_label -> dict data

for row in ws.iter_rows(values_only=True):
    values = tuple(row)
    if not any(v not in (None, "") for v in values):
        continue
    if headers is None:
        joined = " ".join(str(v).lower() for v in values if v)
        if "bps prov" in joined or "bulan sp2d" in joined:
            headers = [str(v).strip() if v is not None else "" for v in values]
        continue

    # Kolom 0=No, 1=BPS, 2=Bulan, 14=TA
    bps_val = str(values[1]).strip().lower() if len(values) > 1 and values[1] else ""
    bulan_val = str(values[2]).strip().lower() if len(values) > 2 and values[2] else ""
    ta_val = values[14] if len(values) > 14 else None

    # Filter: satker target, bulan Mei, TA 2026
    if bps_val not in TARGET_SATKERS:
        continue
    if bulan_val != TARGET_BULAN:
        continue
    if ta_val and int(ta_val) != TARGET_TAHUN:
        continue

    excel_rows[bps_val] = dict(zip(headers, values))

wb.close()

# ================================================================
# BACA SQLITE MonitoringSummary
# ================================================================
satker_codes = ["1300", "1301", "1302", "1303", "1304"]
db_rows = {
    f"bps{row.satker_code}": row
    for row in MonitoringSummary.objects.filter(
        tahun=TARGET_TAHUN, bulan_number=5, satker_code__in=satker_codes
    )
}

# ================================================================
# KOLOM MAPPING (Excel header → SQLite field)
# ================================================================
COLS = [
    ("Realisasi FA 16 Detil Bulan ini (di isi satker)", "fa16_bulan_ini"),
    ("Realisasi Intermilan Bulan ini", "intermilan_bulan_ini"),
    ("Realisasi Intermilan s.d Bulan Ini", "intermilan_sd_bulan_ini"),
    ("Persentase Realisasi Intermilan terhadap FA 16 Detil (Max 100%)", "persen_realisasi"),
    ("Persentase Kelengkapan Dokumen", "persen_kelengkapan_dokumen"),
    ("Persentase SPJ yang sudah di Upload", "persen_spj_upload"),
    ("Persentase dokumen sudah di arsipkan", "persen_arsip"),
    ("% Completed", "percent_completed"),
    ("TA", None),  # Hanya dari Excel
]

print(f"=== PERBANDINGAN EXCEL vs SQLITE: MEI {TARGET_TAHUN} ===\n")
print(f"{'Satker':<12} {'Kolom':<52} {'EXCEL':>18} {'SQLITE':>18} {'MATCH'}")
print("-" * 110)

for satker in TARGET_SATKERS:
    ex = excel_rows.get(satker, {})
    db = db_rows.get(satker)

    if not ex and not db:
        print(f"\n[{satker}] TIDAK ADA di Excel maupun SQLite")
        continue
    elif not ex:
        print(f"\n[{satker}] Ada di SQLite tapi TIDAK ADA di Excel")
        continue
    elif not db:
        print(f"\n[{satker}] Ada di Excel tapi TIDAK ADA di SQLite")
        continue

    print(f"\n[{satker}]")
    for excel_col, db_field in COLS:
        ex_val = ex.get(excel_col, "N/A")
        if db_field is None:
            print(f"  {'':10} {excel_col:<52} {str(ex_val):>18}  {'(hanya Excel)':>18}")
            continue
        db_val = getattr(db, db_field, "N/A")

        # Normalize untuk compare
        try:
            ex_num = float(str(ex_val).replace(",", ".").replace("%", "")) if ex_val not in (None, "", "N/A") else 0.0
            db_num = float(db_val) if db_val not in (None, "", "N/A") else 0.0
            # Persen Excel bisa dalam form 0.xx atau xx.xx → normalize
            if db_field in ("persen_realisasi", "persen_kelengkapan_dokumen", "persen_spj_upload", "persen_arsip", "percent_completed"):
                if 0 < ex_num <= 1:
                    ex_num = round(ex_num * 100, 2)
            match = "OK" if abs(ex_num - db_num) < 0.1 else "!!! BEDA !!!"
        except Exception:
            match = "?"

        print(f"  {'':10} {excel_col:<52} {str(ex_val):>18} {str(db_val):>18}  {match}")

print("\n\n=== RINGKASAN ===")
print(f"Excel rows ditemukan : {list(excel_rows.keys())}")
print(f"SQLite rows ditemukan: {list(db_rows.keys())}")
print(f"Total SQLite all bulan: {MonitoringSummary.objects.count()}")
print(f"Total SQLite Mei 2026 : {MonitoringSummary.objects.filter(tahun=2026, bulan_number=5).count()}")
