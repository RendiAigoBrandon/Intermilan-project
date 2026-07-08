"""Baca sheet Monitoring_Combine dari INTERMILAN.xlsx untuk referensi perbandingan."""
from pathlib import Path
from openpyxl import load_workbook

XLSX_PATH = Path(r"c:\Users\muall\Documents\INTERMILAN PROJECT anti gravity\INTERMILAN PROJECT COPY\data_sources\Database awal\drive-download-20260615T061835Z-3-001\INTERMILAN.xlsx")

print(f"Membuka: {XLSX_PATH}")
wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
print(f"Sheet names: {wb.sheetnames}")

if "Monitoring_Combine" not in wb.sheetnames:
    print("Sheet Monitoring_Combine TIDAK DITEMUKAN.")
    wb.close()
    exit(1)

ws = wb["Monitoring_Combine"]
headers = None
rows_found = []
TARGET_SATKERS = {"bps1300", "bps1301", "bps1302", "bps1303", "bps1304"}

for row in ws.iter_rows(values_only=True):
    values = tuple(row)
    if not any(v not in (None, "") for v in values):
        continue
    if headers is None:
        # Deteksi baris header
        joined = " ".join(str(v).lower() for v in values if v)
        if "bps prov" in joined or "bulan sp2d" in joined:
            headers = [str(v) if v is not None else "" for v in values]
            print(f"\n=== HEADERS ({len(headers)} kolom) ===")
            for i, h in enumerate(headers):
                if h.strip():
                    print(f"  [{i}] {h!r}")
        continue
    # Ambil data
    if not values[0]:
        continue
    satker_val = str(values[0]).strip().lower()
    if any(sk in satker_val for sk in ["bps1300", "bps1301", "bps1302", "bps1303", "bps1304"]):
        bulan_val = str(values[1]).strip().lower() if len(values) > 1 else ""
        if "mei" in bulan_val or "may" in bulan_val:
            rows_found.append(dict(zip(headers, values)))

wb.close()

if rows_found:
    print(f"\n=== DATA MEI 2026 (Monitoring_Combine) — {len(rows_found)} baris ===")
    for row in rows_found:
        print()
        for k, v in row.items():
            if k and str(k).strip():
                print(f"  {k!r}: {v}")
else:
    print("\nTIDAK ADA data Mei untuk bps1300-bps1304 di Monitoring_Combine.")
    print("Mungkin filter bulan perlu disesuaikan.")
    # Tampilkan sample 10 baris pertama agar lihat format datanya
    wb2 = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws2 = wb2["Monitoring_Combine"]
    h2 = None
    count = 0
    print("\n=== SAMPLE 20 BARIS PERTAMA (setelah header) ===")
    for row in ws2.iter_rows(values_only=True):
        values = tuple(row)
        if not any(v not in (None, "") for v in values):
            continue
        if h2 is None:
            joined = " ".join(str(v).lower() for v in values if v)
            if "bps prov" in joined or "bulan sp2d" in joined:
                h2 = [str(v) if v is not None else "" for v in values]
            continue
        if count >= 20:
            break
        if values[0]:
            print(f"  {values[:6]}")
            count += 1
    wb2.close()
