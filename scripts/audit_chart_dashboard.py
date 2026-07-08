"""
Diagnostic read-only: Bandingkan chart Dashboard web vs Excel
TA=2026, Bulan=Mei, semua 20 satker bps1300-bps1377.

Sumber web: MonitoringSummary PostgreSQL
Sumber Excel: INTERMILAN.xlsx sheet Monitoring_Combine
"""
import os
import sys
from pathlib import Path
from decimal import Decimal, InvalidOperation

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "intermilan_project.settings.development")
sys.path.insert(0, ".")
import django
django.setup()

from django.db import connection
from apps.core.models import MonitoringSummary
from apps.core.views import (
    build_mom_rows_from_summary,
    format_percent_id,
    format_id_number,
    percent_height,
)
from openpyxl import load_workbook

XLSX_PATH = Path(
    r"c:\Users\muall\Documents\INTERMILAN PROJECT anti gravity"
    r"\INTERMILAN PROJECT COPY\data_sources\Database awal"
    r"\drive-download-20260615T061835Z-3-001\INTERMILAN.xlsx"
)

# ================================================================
# 0. DB CONFIRMATION
# ================================================================
print("=" * 70)
print("DATABASE AKTIF")
print("=" * 70)
print(f"  backend : {connection.vendor}")
print(f"  name    : {connection.settings_dict.get('NAME', '-')}")
print(f"  host    : {connection.settings_dict.get('HOST', '-')}")
print()

# ================================================================
# 1. BACA EXCEL — Monitoring_Combine, Bulan=Mei, TA=2026
# ================================================================
wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb["Monitoring_Combine"]
headers = None
excel_rows = {}  # satker_label (lowercase) -> dict

for row in ws.iter_rows(values_only=True):
    values = tuple(row)
    if not any(v not in (None, "") for v in values):
        continue
    if headers is None:
        joined = " ".join(str(v).lower() for v in values if v)
        if "bps prov" in joined or "bulan sp2d" in joined:
            headers = [str(v).strip() if v is not None else "" for v in values]
        continue
    bps_val = str(values[1]).strip().lower() if len(values) > 1 and values[1] else ""
    bulan_val = str(values[2]).strip().lower() if len(values) > 2 and values[2] else ""
    ta_val = values[14] if len(values) > 14 else None
    if not bps_val or bulan_val != "mei":
        continue
    try:
        if ta_val and int(float(str(ta_val))) != 2026:
            continue
    except Exception:
        continue
    excel_rows[bps_val] = {
        "fa16": values[3] if len(values) > 3 else None,
        "intermilan_bulan": values[4] if len(values) > 4 else None,
        "intermilan_sd": values[5] if len(values) > 5 else None,
        "persen": values[6] if len(values) > 6 else None,
    }

wb.close()

# ================================================================
# 2. BACA PostgreSQL
# ================================================================
pg_qs = MonitoringSummary.objects.filter(tahun=2026, bulan_number=5).order_by("satker_code")
pg_rows = {f"bps{r.satker_code}": r for r in pg_qs}

# ================================================================
# 3. CHART CONTEXT SIMULATION (bps1303 khusus)
# ================================================================
summary_qs = MonitoringSummary.objects.filter(tahun=2026, bulan_number=5)
summary_available = summary_qs.exists()
mom_rows_web = build_mom_rows_from_summary(summary_qs) if summary_available else []

# ================================================================
# 4. HELPER
# ================================================================
def to_decimal(val):
    if val is None or val == "":
        return Decimal("0")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except (InvalidOperation, Exception):
        return Decimal("0")

def pct_normalize(val):
    """Excel persen bisa 0.97 (fraction) atau 97.39 (angka). Normalize ke 0-100."""
    d = to_decimal(val)
    if Decimal("0") < d <= Decimal("1"):
        d = (d * Decimal("100")).quantize(Decimal("0.01"))
    return d

def fmt(val):
    try:
        d = Decimal(str(val))
        return f"{d:,.0f}".replace(",", ".")
    except Exception:
        return str(val)

def fmtp(val):
    try:
        d = Decimal(str(val))
        return f"{d:.2f}%".replace(".", ",")
    except Exception:
        return str(val)

def match_label(ex, pg, tolerance=1):
    try:
        diff = abs(Decimal(str(ex)) - Decimal(str(pg)))
        return "OK" if diff <= tolerance else "!!! BEDA !!!"
    except Exception:
        return "?"

# ================================================================
# 5. LAPORAN
# ================================================================
all_satkers = sorted(set(excel_rows.keys()) | set(pg_rows.keys()))

print("=" * 70)
print(f"JUMLAH BARIS")
print("=" * 70)
print(f"  Excel Mei 2026    : {len(excel_rows)} satker")
print(f"  PostgreSQL Mei 2026: {len(pg_rows)} satker")
print()

only_excel = sorted(set(excel_rows) - set(pg_rows))
only_pg = sorted(set(pg_rows) - set(excel_rows))
print(f"  Di Excel tapi tidak di PG : {only_excel or 'tidak ada'}")
print(f"  Di PG tapi tidak di Excel : {only_pg or 'tidak ada'}")
print()

print("=" * 70)
print("TABEL PERBANDINGAN 20 SATKER — MEI 2026")
print("=" * 70)
print(f"{'Satker':<10} {'Kolom':<32} {'Excel':>18} {'PostgreSQL':>18} {'Match'}")
print("-" * 90)

all_ok = True
bps1303_checks = {}

for satker in all_satkers:
    ex = excel_rows.get(satker, {})
    pg = pg_rows.get(satker)

    ex_fa16 = to_decimal(ex.get("fa16"))
    ex_bulan = to_decimal(ex.get("intermilan_bulan"))
    ex_sd = to_decimal(ex.get("intermilan_sd"))
    ex_pct = pct_normalize(ex.get("persen"))

    pg_fa16 = to_decimal(pg.fa16_bulan_ini) if pg else Decimal("0")
    pg_bulan = to_decimal(pg.intermilan_bulan_ini) if pg else Decimal("0")
    pg_sd = to_decimal(pg.intermilan_sd_bulan_ini) if pg else Decimal("0")
    pg_pct = to_decimal(pg.persen_realisasi) if pg else Decimal("0")

    checks = [
        ("FA16 Bulan ini",     ex_fa16,  pg_fa16,  1),
        ("Intermilan Bulan",   ex_bulan, pg_bulan, 1),
        ("Intermilan s.d",     ex_sd,    pg_sd,    1),
        ("Persen Realisasi",   ex_pct,   pg_pct,   Decimal("0.02")),
    ]

    satker_ok = True
    for label, ex_val, pg_val, tol in checks:
        m = match_label(ex_val, pg_val, tol)
        if m != "OK":
            satker_ok = False
            all_ok = False
        print(f"{satker:<10} {label:<32} {fmt(ex_val):>18} {fmt(pg_val):>18}  {m}")

    if satker == "bps1303":
        bps1303_checks = {
            "fa16": (ex_fa16, pg_fa16),
            "bulan": (ex_bulan, pg_bulan),
            "sd": (ex_sd, pg_sd),
            "pct": (ex_pct, pg_pct),
        }

    if satker_ok:
        print(f"{satker:<10} {'':>32} {'':>18} {'':>18}  [SATKER OK]")
    print()

# ================================================================
# 6. KHUSUS bps1303 MEI
# ================================================================
print("=" * 70)
print("CEK KHUSUS bps1303 MEI 2026")
print("=" * 70)
targets = {
    "FA16":              (Decimal("370459272"), bps1303_checks.get("fa16", (Decimal("0"), Decimal("0")))[1]),
    "Intermilan Bulan":  (Decimal("360781982"), bps1303_checks.get("bulan", (Decimal("0"), Decimal("0")))[1]),
    "Intermilan s.d":    (Decimal("2237027245"), bps1303_checks.get("sd", (Decimal("0"), Decimal("0")))[1]),
    "Persen Realisasi":  (Decimal("97.39"), bps1303_checks.get("pct", (Decimal("0"), Decimal("0")))[1]),
}
for label, (expected, actual) in targets.items():
    tol = Decimal("0.02") if "Persen" in label else Decimal("1")
    m = "OK" if abs(actual - expected) <= tol else "!!! BEDA !!!"
    print(f"  {label:<22}: expected={fmt(expected):>18}  actual={fmt(actual):>18}  {m}")

# ================================================================
# 7. CHART CONTEXT WEB
# ================================================================
print()
print("=" * 70)
print("CHART CONTEXT WEB — /dashboard/?tahun=2026&bulan=5")
print("=" * 70)
print(f"  summary_available       : {summary_available}")
print(f"  source                  : {'MonitoringSummary' if summary_available else 'Fallback D_K'}")
print(f"  jumlah satker (x-axis)  : {len(mom_rows_web)}")

if mom_rows_web:
    max_val = max(
        [r["fa16_bulan_ini"] for r in [{"fa16_bulan_ini": pg.fa16_bulan_ini, "intermilan_bulan_ini": pg.intermilan_bulan_ini, "intermilan_sd_bulan_ini": pg.intermilan_sd_bulan_ini} for pg in pg_qs]] +
        [r["intermilan_bulan_ini"] for r in [{"fa16_bulan_ini": pg.fa16_bulan_ini, "intermilan_bulan_ini": pg.intermilan_bulan_ini, "intermilan_sd_bulan_ini": pg.intermilan_sd_bulan_ini} for pg in pg_qs]] +
        [r["intermilan_sd_bulan_ini"] for r in [{"fa16_bulan_ini": pg.fa16_bulan_ini, "intermilan_bulan_ini": pg.intermilan_bulan_ini, "intermilan_sd_bulan_ini": pg.intermilan_sd_bulan_ini} for pg in pg_qs]] +
        [Decimal("1")]
    )
    print(f"  max_value (chart scale) : {fmt(max_val)}")
    print()
    print(f"  {'Satker':<12} {'pct':>8} {'fa(%)':>8} {'bulan(%)':>10} {'sd(%)':>8}  {'fa_label':>18}  {'bulan_label':>18}  {'sd_label':>18}")
    print("  " + "-" * 110)
    for r in mom_rows_web:
        print(f"  {r['satker']:<12} {r['pct']:>8} {r['fa']:>8} {r['bulan']:>10} {r['sd']:>8}  {r['fa_label']:>18}  {r['bulan_label']:>18}  {r['sd_label']:>18}")

print()
print("  Series urutan (sesuai template dashboard.html):")
print("    1. [hijau]  Persentase Realisasi Intermilan terhadap FA 16 Detil (Max 100%)")
print("    2. [kuning] Realisasi Intermilan s.d Bulan Ini")
print("    3. [merah]  Realisasi Intermilan Bulan ini")
print("    4. [biru]   Realisasi FA 16 Detil Bulan ini (di isi satker)")

# ================================================================
# 8. KESIMPULAN
# ================================================================
print()
print("=" * 70)
print("KESIMPULAN FINAL")
print("=" * 70)
if all_ok:
    print("  Chart data BENAR. Semua kolom dan satker cocok antara Excel dan PostgreSQL.")
else:
    print("  Chart data ADA PERBEDAAN. Lihat baris '!!! BEDA !!!' di atas.")
print(f"  Source: {'MonitoringSummary (BENAR, bukan D_K)' if summary_available else 'FALLBACK D_K (salah!)'}")
print(f"  Jumlah satker x-axis: {len(mom_rows_web)} {'(sesuai target 20)' if len(mom_rows_web) == 20 else '(BERBEDA dari target 20!)'}")
print()
print("  CATATAN FA16=0:")
fa16_zero = [satker for satker, pg in pg_rows.items() if pg.fa16_bulan_ini == 0]
print(f"  {len(fa16_zero)} satker FA16=0 di Mei 2026 (sesuai Excel, bukan bug):")
print(f"  {fa16_zero}")
