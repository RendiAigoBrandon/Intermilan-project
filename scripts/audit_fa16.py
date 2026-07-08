"""
Audit FA16: bandingkan fa16_bulan_ini antara Excel dan PostgreSQL
untuk semua satker, bulan Mei, tahun 2026.
"""
import os
import sys
from pathlib import Path
from decimal import Decimal

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "intermilan_project.settings.development")
sys.path.insert(0, ".")
import django
django.setup()

from apps.core.models import MonitoringSummary
from openpyxl import load_workbook

XLSX_PATH = Path(
    r"c:\Users\muall\Documents\INTERMILAN PROJECT anti gravity"
    r"\INTERMILAN PROJECT COPY\data_sources\Database awal"
    r"\drive-download-20260615T061835Z-3-001\INTERMILAN.xlsx"
)

# ================================================================
# BACA EXCEL — semua satker, bulan Mei, TA 2026
# ================================================================
wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb["Monitoring_Combine"]
headers = None
excel_fa16 = {}  # satker_label -> fa16 value

for row in ws.iter_rows(values_only=True):
    values = tuple(row)
    if not any(v not in (None, "") for v in values):
        continue
    if headers is None:
        joined = " ".join(str(v).lower() for v in values if v)
        if "bps prov" in joined or "bulan sp2d" in joined:
            headers = [str(v).strip() if v is not None else "" for v in values]
        continue
    bps_val = str(values[1]).strip().lower() if len(values) > 1 and values[1] else ""
    bulan_val = str(values[2]).strip().lower() if len(values) > 2 and values[2] else ""
    ta_val = values[14] if len(values) > 14 else None
    if not bps_val or bulan_val != "mei":
        continue
    if ta_val and str(int(float(ta_val))) != "2026":
        continue
    fa16_val = values[3] if len(values) > 3 else None
    excel_fa16[bps_val] = fa16_val if fa16_val is not None else 0

wb.close()

# ================================================================
# BACA PostgreSQL
# ================================================================
pg_rows = {
    f"bps{r.satker_code}": r
    for r in MonitoringSummary.objects.filter(tahun=2026, bulan_number=5).order_by("satker_code")
}

# ================================================================
# PERBANDINGAN
# ================================================================
print("=== AUDIT FA16 MEI 2026: EXCEL vs POSTGRESQL ===\n")
print(f"{'Satker':<12} {'Excel FA16':>18} {'PG FA16':>18} {'Match'}")
print("-" * 60)

fa16_zero_excel = []
fa16_zero_pg = []
mismatch = []

all_satkers = sorted(set(excel_fa16.keys()) | set(pg_rows.keys()))
for satker in all_satkers:
    ex_val = excel_fa16.get(satker, "TIDAK ADA DI EXCEL")
    pg_row = pg_rows.get(satker)
    pg_val = float(pg_row.fa16_bulan_ini) if pg_row else "TIDAK ADA DI PG"

    # Check zero
    if isinstance(ex_val, (int, float, Decimal)) and float(ex_val) == 0:
        fa16_zero_excel.append(satker)
    elif ex_val in (None, 0):
        fa16_zero_excel.append(satker)

    if isinstance(pg_val, float) and pg_val == 0:
        fa16_zero_pg.append(satker)

    # Match check
    try:
        ex_num = float(ex_val) if ex_val not in ("TIDAK ADA DI EXCEL", None) else None
        pg_num = float(pg_val) if isinstance(pg_val, float) else None
        if ex_num is None or pg_num is None:
            match = "N/A"
        elif abs(ex_num - pg_num) < 1:
            match = "OK"
        else:
            match = "!!! BEDA !!!"
            mismatch.append(satker)
    except Exception:
        match = "?"

    print(f"{satker:<12} {str(ex_val):>18} {str(pg_val):>18} {match}")

print()
print("=== RINGKASAN FA16 ===")
print(f"FA16=0 di Excel Mei 2026 : {len(fa16_zero_excel)} satker")
print(f"  -> {fa16_zero_excel}")
print(f"FA16=0 di PG Mei 2026    : {len(fa16_zero_pg)} satker")
print(f"  -> {fa16_zero_pg}")
print(f"Mismatch Excel vs PG     : {len(mismatch)} satker")
if mismatch:
    print(f"  -> {mismatch}")
    print("KESIMPULAN: Ada perbedaan mapping FA16 antara Excel dan PostgreSQL.")
else:
    daftar_sama = set(fa16_zero_excel) == set(fa16_zero_pg)
    if daftar_sama:
        print("KESIMPULAN: FA16=0 SESUAI EXCEL. Bukan bug mapping. Nilai 0 memang dari sumber Excel.")
    else:
        print("KESIMPULAN: Daftar satker FA16=0 BEDA antara Excel dan PG. Perlu investigasi.")
