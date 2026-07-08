from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from apps.core.import_utils import clean_text, dict_from_headers, parse_date, parse_decimal, parse_month, pick
from apps.core.models import MonitoringSummary

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


class Command(BaseCommand):
    help = "Import INTERMILAN.xlsx sheet Monitoring_Combine sebagai baseline MonitoringSummary."

    def add_arguments(self, parser):
        parser.add_argument("--path", required=True, help="Path INTERMILAN.xlsx atau folder Database awal.")
        parser.add_argument("--commit", action="store_true", help="Jalankan import asli. Default dry-run.")
        parser.add_argument("--limit-rows", type=int, default=0, help="Batasi baris untuk audit cepat.")

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError("openpyxl belum tersedia.")
        workbook_path = self.find_workbook(Path(options["path"]))
        self.commit = options["commit"]
        self.limit_rows = options["limit_rows"]
        self.stdout.write(self.style.WARNING("Mode: IMPORT ASLI") if self.commit else self.style.WARNING("Mode: DRY-RUN"))
        self.stdout.write(f"Workbook: {workbook_path}")

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            if "Monitoring_Combine" not in wb.sheetnames:
                raise CommandError("Sheet Monitoring_Combine tidak ditemukan.")
            stats = self.import_sheet(wb["Monitoring_Combine"])
        finally:
            wb.close()

        self.stdout.write(self.style.HTTP_INFO("Ringkasan import MonitoringSummary"))
        self.stdout.write(
            f"read={stats['read']}, success={stats['success']}, skip={stats['skip']}, "
            f"updated={stats['updated']}, failed={stats['failed']}"
        )

    def find_workbook(self, path: Path) -> Path:
        if not path.exists():
            raise CommandError(f"Path tidak ditemukan: {path}")
        if path.is_file():
            if path.name.upper() != "INTERMILAN.XLSX":
                raise CommandError("File harus INTERMILAN.xlsx.")
            return path
        direct = path / "INTERMILAN.xlsx"
        if direct.exists():
            return direct
        matches = sorted(path.rglob("INTERMILAN.xlsx"))
        if not matches:
            raise CommandError(f"INTERMILAN.xlsx tidak ditemukan di: {path}")
        return matches[0]

    def import_sheet(self, ws):
        headers = None
        stats = {"read": 0, "success": 0, "skip": 0, "updated": 0, "failed": 0}
        for row in ws.iter_rows(values_only=True):
            values = tuple(row)
            if not any(v not in (None, "") for v in values):
                continue
            if headers is None:
                normalized = [clean_text(v).lower() for v in values]
                if "bps prov/kab/kota" in " ".join(normalized) and "bulan sp2d" in " ".join(normalized):
                    headers = list(values)
                continue
            if self.limit_rows and stats["read"] >= self.limit_rows:
                break
            stats["read"] += 1
            row_data = dict_from_headers(headers, values)
            satker_label = clean_text(pick(row_data, "bps prov kab kota"))
            bulan = clean_text(pick(row_data, "bulan sp2d"))
            tahun_raw = pick(row_data, "ta")
            bulan_number = parse_month(bulan)
            tahun = int(tahun_raw) if tahun_raw not in (None, "") else 0
            if not satker_label or not bulan_number or not tahun:
                stats["skip"] += 1
                continue
            defaults = {
                "satker_label": satker_label,
                "bulan": bulan,
                "fa16_bulan_ini": parse_decimal(pick(row_data, "realisasi fa 16 detil bulan ini di isi satker")),
                "intermilan_bulan_ini": parse_decimal(pick(row_data, "realisasi intermilan bulan ini")),
                "intermilan_sd_bulan_ini": parse_decimal(pick(row_data, "realisasi intermilan s d bulan ini")),
                "persen_realisasi": parse_percent(pick(row_data, "persentase realisasi intermilan terhadap fa 16 detil max 100%")),
                "persen_kelengkapan_dokumen": parse_percent(pick(row_data, "persentase kelengkapan dokumen")),
                "persen_spj_upload": parse_percent(pick(row_data, "persentase spj yang sudah di upload")),
                "persen_arsip": parse_percent(pick(row_data, "persentase dokumen sudah di arsipkan")),
                "deadline": parse_date(pick(row_data, "deadline")),
                "status": clean_text(pick(row_data, "status")),
                "percent_completed": parse_percent(pick(row_data, "% completed")),
                "bar": clean_text(pick(row_data, "bar")),
                "source": MonitoringSummary.Source.EXCEL_SEED,
            }
            if not self.commit:
                stats["success"] += 1
                continue
            _, created = MonitoringSummary.objects.update_or_create(
                satker_code=satker_code_from_label(satker_label),
                bulan_number=bulan_number,
                tahun=tahun,
                defaults=defaults,
            )
            stats["success"] += 1
            if not created:
                stats["updated"] += 1
        return stats


def satker_code_from_label(value: str) -> str:
    text = clean_text(value).lower()
    return text.removeprefix("bps").strip()


def parse_percent(value):
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        decimal = Decimal(str(value))
        if Decimal("0") <= decimal <= Decimal("1"):
            decimal *= Decimal("100")
        return decimal.quantize(Decimal("0.01"))
    text = clean_text(value).replace("%", "").strip()
    if not text:
        return Decimal("0")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0")
