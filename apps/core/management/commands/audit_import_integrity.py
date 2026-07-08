from __future__ import annotations

import re
import sqlite3
import warnings
from collections import Counter
from datetime import datetime
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import connection
from django.db.models import Count
from django.utils import timezone

from apps.dk.models import MasterAkun, TransactionDetail
from apps.documents.models import ChecklistStatus, ChecklistTemplate, DocumentDriveLink
from apps.drpp.models import DRPPItem, DRPPMatch, DRPPUpload
from apps.sp2d.models import SP2DImportBatch, SP2DRaw

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


class Command(BaseCommand):
    help = "Audit read-only integritas data hasil import INTERMILAN Tahap 2."

    def add_arguments(self, parser):
        parser.add_argument("--legacy-sqlite-path", default="", help="Path SQLite legacy untuk membedakan source legacy vs tambahan Excel.")
        parser.add_argument("--excel-path", default="", help="Path folder/file Excel seed untuk audit warning tanggal.")
        parser.add_argument("--output", default="", help="Path file Markdown report.")
        parser.add_argument("--max-samples", type=int, default=30, help="Jumlah sample maksimum per temuan.")

    def handle(self, *args, **options):
        self.max_samples = options["max_samples"]
        self.legacy_path = Path(options["legacy_sqlite_path"]) if options["legacy_sqlite_path"] else None
        self.excel_path = Path(options["excel_path"]) if options["excel_path"] else None

        report = self.build_report()
        self.stdout.write(report)

        output = options["output"]
        if output:
            output_path = Path(output)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(report, encoding="utf-8")
            self.stdout.write(self.style.SUCCESS(f"\nReport ditulis ke: {output_path}"))

    def build_report(self) -> str:
        generated_at = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S %Z")
        lines = [
            "# Audit Integritas Data Import INTERMILAN",
            "",
            f"- Waktu audit: {generated_at}",
            "- Mode: read-only, tidak ada hapus/replace/drop.",
            "",
        ]

        lines += self.audit_counts()
        lines += self.audit_transaction_details()
        lines += self.audit_master_akun()
        lines += self.audit_checklist_template()
        lines += self.audit_checklist_status()
        lines += self.audit_drive_links()
        lines += self.audit_sp2d()
        lines += self.audit_excel_date_warnings()
        lines += self.audit_dummy_data()
        lines += self.recommendations()
        return "\n".join(lines) + "\n"

    def audit_counts(self) -> list[str]:
        counts = [
            ("SP2D batch", SP2DImportBatch.objects.count()),
            ("SP2D raw", SP2DRaw.objects.count()),
            ("D_K / TransactionDetail", TransactionDetail.objects.count()),
            ("Master Akun", MasterAkun.objects.count()),
            ("Checklist template", ChecklistTemplate.objects.count()),
            ("Checklist status", ChecklistStatus.objects.count()),
            ("DocumentDriveLink", DocumentDriveLink.objects.count()),
            ("DRPP upload", DRPPUpload.objects.count()),
            ("DRPP item", DRPPItem.objects.count()),
            ("DRPP match", DRPPMatch.objects.count()),
        ]
        lines = ["## 1. Ringkasan Jumlah Data", ""]
        for label, value in counts:
            lines.append(f"- {label}: {value:,}".replace(",", "."))
        lines.append("")
        return lines

    def audit_transaction_details(self) -> list[str]:
        total = TransactionDetail.objects.count()
        legacy_ids = self.read_legacy_ids("transaksi_detail")
        legacy_count = TransactionDetail.objects.filter(id__in=legacy_ids).count() if legacy_ids else 0
        extra_count = total - legacy_count if legacy_ids else 0
        duplicate_groups = list(
            TransactionDetail.objects.values(
                "satker_code", "nomor_spm", "no_kuitansi", "no_drpp", "akun", "nilai_bruto", "nilai_netto"
            )
            .annotate(total=Count("id"))
            .filter(total__gt=1)
            .order_by("-total")[: self.max_samples]
        )
        duplicate_total = (
            TransactionDetail.objects.values(
                "satker_code", "nomor_spm", "no_kuitansi", "no_drpp", "akun", "nilai_bruto", "nilai_netto"
            )
            .annotate(total=Count("id"))
            .filter(total__gt=1)
            .count()
        )
        blank_key_count = TransactionDetail.objects.filter(nomor_spm="", no_kuitansi="", no_drpp="").count()

        lines = [
            "## 2. Audit D_K / TransactionDetail",
            "",
            f"- Total Django saat ini: {total:,}".replace(",", "."),
            f"- Baris yang ID-nya cocok dengan SQLite legacy: {legacy_count:,}".replace(",", "."),
            f"- Baris tambahan non-legacy, kemungkinan dari Excel/manual: {extra_count:,}".replace(",", "."),
            f"- Baris tanpa nomor SPM/kuitansi/DRPP: {blank_key_count:,}".replace(",", "."),
            f"- Grup duplikat kandidat berdasarkan satker+SPM+kuitansi+DRPP+akun+nilai bruto/netto: {duplicate_total:,}".replace(",", "."),
            "",
            "Sample grup duplikat kandidat:",
        ]
        if duplicate_groups:
            for item in duplicate_groups:
                lines.append(
                    f"- {item['satker_code']} | SPM {item['nomor_spm']} | KW {item['no_kuitansi']} | "
                    f"DRPP {item['no_drpp']} | akun {item['akun']} | bruto {item['nilai_bruto']} | "
                    f"netto {item['nilai_netto']} -> {item['total']} baris"
                )
        else:
            lines.append("- Tidak ditemukan.")
        lines.append("")
        return lines

    def audit_master_akun(self) -> list[str]:
        legacy_codes = self.read_legacy_master_codes()
        codes = list(MasterAkun.objects.values_list("kode", flat=True))
        counter = Counter(codes)
        duplicates = [code for code, count in counter.items() if count > 1]
        empty = [code for code in codes if not str(code).strip()]
        invalid = [code for code in codes if code and not re.fullmatch(r"[0-9Xx]+", code)]
        additional = sorted(set(codes) - legacy_codes) if legacy_codes else []
        category_blank = list(MasterAkun.objects.filter(kategori="").values_list("kode", "nama_akun")[: self.max_samples])

        lines = [
            "## 3. Audit Master Akun",
            "",
            f"- Total Master Akun: {len(codes)}",
            f"- Kode tambahan dibanding SQLite legacy: {', '.join(additional) if additional else '-'}",
            f"- Kode kosong: {len(empty)}",
            f"- Kode invalid format: {', '.join(invalid[:self.max_samples]) if invalid else '-'}",
            f"- Duplikat kode: {', '.join(duplicates) if duplicates else '-'}",
            f"- Kategori kosong: {len(category_blank)} sample",
        ]
        for kode, nama in category_blank:
            lines.append(f"  - {kode}: {nama}")
        lines.append("")
        return lines

    def audit_checklist_template(self) -> list[str]:
        total = ChecklistTemplate.objects.count()
        distinct_names = ChecklistTemplate.objects.values("nama_dokumen").distinct().count()
        duplicate_names = list(
            ChecklistTemplate.objects.values("nama_dokumen")
            .annotate(total=Count("id"))
            .filter(total__gt=1)
            .order_by("-total", "nama_dokumen")[: self.max_samples]
        )
        duplicate_name_category = (
            ChecklistTemplate.objects.values("nama_dokumen", "kategori")
            .annotate(total=Count("id"))
            .filter(total__gt=1)
            .count()
        )
        names = list(ChecklistTemplate.objects.values_list("nama_dokumen", flat=True).distinct().order_by("nama_dokumen"))
        legacy_template_count = self.read_legacy_count("checklist_template")

        lines = [
            "## 4. Audit ChecklistTemplate",
            "",
            f"- Total ChecklistTemplate Django: {total:,}".replace(",", "."),
            f"- Total template di SQLite legacy: {legacy_template_count if legacy_template_count is not None else '-'}",
            f"- Distinct nama_dokumen: {distinct_names}",
            f"- Duplikat exact nama_dokumen+kategori: {duplicate_name_category}",
            "",
            "Distinct nama_dokumen:",
        ]
        for name in names:
            lines.append(f"- {name}")
        lines += ["", "Sample nama_dokumen yang berulang lintas kategori/pattern:"]
        if duplicate_names:
            for item in duplicate_names:
                lines.append(f"- {item['nama_dokumen']}: {item['total']} template")
        else:
            lines.append("- Tidak ditemukan.")
        lines += [
            "",
            "Catatan audit: jumlah 601 berasal dari tabel `checklist_template` SQLite legacy yang menyimpan kombinasi pattern/kategori, bukan dari `checklist_status`. Secara master UI, jumlah ideal kemungkinan cukup sebesar distinct nama_dokumen.",
            "",
        ]
        return lines

    def audit_checklist_status(self) -> list[str]:
        total = ChecklistStatus.objects.count()
        transaction_total = TransactionDetail.objects.count()
        avg = total / transaction_total if transaction_total else 0
        orphan = self.count_orphans("documents_checkliststatus", "transaction_detail_id", "dk_transactiondetail")
        duplicate_pairs = (
            ChecklistStatus.objects.values("transaction_detail_id", "nama_dokumen")
            .annotate(total=Count("id"))
            .filter(total__gt=1)
            .count()
        )
        status_breakdown = list(ChecklistStatus.objects.values("status").annotate(total=Count("id")).order_by("status"))
        lines = [
            "## 5. Audit ChecklistStatus",
            "",
            f"- Total ChecklistStatus: {total:,}".replace(",", "."),
            f"- Total TransactionDetail: {transaction_total:,}".replace(",", "."),
            f"- Rata-rata checklist per transaksi: {avg:.2f}",
            f"- Orphan transaction_detail: {orphan}",
            f"- Duplikat transaction_detail + nama_dokumen: {duplicate_pairs}",
            "- Breakdown status:",
        ]
        for item in status_breakdown:
            lines.append(f"  - {item['status']}: {item['total']:,}".replace(",", "."))
        lines.append("")
        return lines

    def audit_drive_links(self) -> list[str]:
        total = DocumentDriveLink.objects.count()
        empty_url = DocumentDriveLink.objects.filter(google_drive_url="").count()
        empty_name = DocumentDriveLink.objects.filter(nama_file="").count()
        matched = DocumentDriveLink.objects.filter(transaction_detail__isnull=False).count()
        unmatched = total - matched
        invalid = list(
            DocumentDriveLink.objects.exclude(google_drive_url__regex=r"^https?://(drive|docs)\.google\.com/")
            .values_list("id", "satker_code", "nomor_spm", "google_drive_url")[: self.max_samples]
        )
        duplicate_urls = list(
            DocumentDriveLink.objects.values("google_drive_url")
            .annotate(total=Count("id"))
            .filter(total__gt=1)
            .order_by("-total")[: self.max_samples]
        )

        lines = [
            "## 6. Audit DocumentDriveLink",
            "",
            f"- Total DocumentDriveLink: {total:,}".replace(",", "."),
            f"- Matched ke TransactionDetail: {matched:,}".replace(",", "."),
            f"- Belum matched: {unmatched:,}".replace(",", "."),
            f"- URL kosong: {empty_url}",
            f"- Nama file/key kosong: {empty_name}",
            f"- URL invalid format Google Drive/Docs: {len(invalid)} sample",
        ]
        for item in invalid:
            lines.append(f"  - id={item[0]} satker={item[1]} spm={item[2]} url={item[3]}")
        lines.append(f"- Sample link duplikat: {len(duplicate_urls)} grup")
        for item in duplicate_urls:
            lines.append(f"  - {item['google_drive_url']} -> {item['total']} baris")
        lines.append("")
        return lines

    def audit_sp2d(self) -> list[str]:
        total = SP2DRaw.objects.count()
        duplicate_no_sp2d = (
            SP2DRaw.objects.exclude(no_sp2d="")
            .values("no_sp2d")
            .annotate(total=Count("id"))
            .filter(total__gt=1)
            .count()
        )
        blank_no_sp2d = SP2DRaw.objects.filter(no_sp2d="").count()
        blank_spm = SP2DRaw.objects.filter(nomor_spm_extracted="").count()
        invalid_value = SP2DRaw.objects.filter(nilai_sp2d__lte=0).count()
        blank_satker = SP2DRaw.objects.filter(satker_code="").count()
        null_dates = SP2DRaw.objects.filter(tgl_sp2d__isnull=True).count()
        sample_duplicates = list(
            SP2DRaw.objects.exclude(no_sp2d="")
            .values("no_sp2d")
            .annotate(total=Count("id"))
            .filter(total__gt=1)
            .order_by("-total")[: self.max_samples]
        )

        lines = [
            "## 7. Audit SP2D",
            "",
            f"- Total SP2DRaw: {total:,}".replace(",", "."),
            f"- Grup duplikat no_sp2d non-kosong: {duplicate_no_sp2d}",
            f"- no_sp2d kosong: {blank_no_sp2d}",
            f"- nomor_spm_extracted kosong: {blank_spm}",
            f"- nilai_sp2d <= 0: {invalid_value}",
            f"- tanggal SP2D null: {null_dates}",
            f"- satker kosong: {blank_satker}",
            "- Sample duplikat no_sp2d:",
        ]
        if sample_duplicates:
            for item in sample_duplicates:
                lines.append(f"  - {item['no_sp2d']}: {item['total']} baris")
        else:
            lines.append("  - Tidak ditemukan.")
        lines.append("")
        return lines

    def audit_excel_date_warnings(self) -> list[str]:
        lines = ["## 8. Audit Warning Tanggal Excel", ""]
        if not self.excel_path or not self.excel_path.exists() or load_workbook is None:
            lines += ["- Excel path tidak diberikan atau openpyxl tidak tersedia.", ""]
            return lines

        files = [self.excel_path] if self.excel_path.is_file() else sorted(self.excel_path.glob("*.xlsx"))
        warnings_found = []
        for file_path in files:
            if file_path.name.startswith("~$"):
                continue
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                wb = load_workbook(file_path, read_only=True, data_only=True)
                try:
                    for ws in wb.worksheets:
                        if ws.title not in {"D_K", "Upload", "Data_Integrasi", "Dashboard", "Monitoring_Combine"}:
                            continue
                        with warnings.catch_warnings(record=True) as sheet_caught:
                            warnings.simplefilter("always")
                            for _ in ws.iter_rows(values_only=True):
                                pass
                            for warning in sheet_caught:
                                warnings_found.append((file_path.name, ws.title, str(warning.message)))
                    for warning in caught:
                        warnings_found.append((file_path.name, "-", str(warning.message)))
                finally:
                    wb.close()

        if warnings_found:
            for filename, sheet, message in warnings_found[: self.max_samples]:
                cell = self.extract_cell(message)
                column_hint = self.column_hint(cell)
                lines.append(f"- File `{filename}`, sheet `{sheet}`, cell `{cell or '-'}`, kolom `{column_hint}`: {message}")
            if len(warnings_found) > self.max_samples:
                lines.append(f"- ... {len(warnings_found) - self.max_samples} warning lain dipotong dari laporan.")
            lines.append("")
            lines.append("Dampak penyimpanan: nilai tanggal yang tidak bisa dibaca openpyxl masuk sebagai error/None pada parser tanggal, sehingga field seperti `tanggal_spm` dapat tersimpan null untuk baris terkait. Tidak ada drop/hapus data.")
        else:
            lines.append("- Tidak ada warning tanggal Excel saat audit.")
        lines.append("")
        return lines

    def audit_dummy_data(self) -> list[str]:
        patterns = ["dummy", "sample", "test", "contoh"]
        lines = ["## 9. Audit Data Dummy / Sample", ""]
        findings = []
        for pattern in patterns:
            findings.append(("TransactionDetail.deskripsi", pattern, TransactionDetail.objects.filter(deskripsi__icontains=pattern).count()))
            findings.append(("SP2DRaw.deskripsi", pattern, SP2DRaw.objects.filter(deskripsi__icontains=pattern).count()))
            findings.append(("DocumentDriveLink.nama_file", pattern, DocumentDriveLink.objects.filter(nama_file__icontains=pattern).count()))
            findings.append(("MasterAkun.nama_akun", pattern, MasterAkun.objects.filter(nama_akun__icontains=pattern).count()))
        nonzero = [item for item in findings if item[2]]
        if nonzero:
            for field, pattern, count in nonzero:
                lines.append(f"- `{field}` mengandung `{pattern}`: {count}")
        else:
            lines.append("- Tidak ditemukan keyword dummy/sample/test/contoh pada field utama yang diaudit.")
        weird_spm = list(
            TransactionDetail.objects.exclude(nomor_spm="")
            .exclude(nomor_spm__regex=r"^[0-9A-Za-z/.\-]+$")
            .values_list("id", "satker_code", "nomor_spm")[: self.max_samples]
        )
        lines.append(f"- Nomor SPM dengan karakter tidak lazim: {len(weird_spm)} sample")
        for item in weird_spm:
            lines.append(f"  - id={item[0]} satker={item[1]} nomor_spm={item[2]}")
        lines.append("")
        return lines

    def recommendations(self) -> list[str]:
        return [
            "## 10. Rekomendasi Perbaikan (Belum Dieksekusi)",
            "",
            "1. Tandai sumber import pada `TransactionDetail` di migration berikutnya bila ingin membedakan legacy vs Excel tanpa membaca SQLite lama.",
            "2. Review `ChecklistTemplate` 601 baris: kemungkinan perlu normalisasi menjadi master distinct nama_dokumen + mapping pattern terpisah.",
            "3. Review D_K tambahan Excel non-legacy dan grup duplikat kandidat sebelum memutuskan merge/delete.",
            "4. Review `DocumentDriveLink` yang belum matched ke transaksi; beberapa link mungkin hanya arsip umum atau belum punya D_K padanan.",
            "5. Review warning tanggal Excel pada workbook terkait sebelum memakai tanggal tersebut untuk laporan resmi.",
            "",
        ]

    def read_legacy_ids(self, table: str) -> set[int]:
        if not self.legacy_path or not self.legacy_path.exists():
            return set()
        conn = sqlite3.connect(self.legacy_path)
        try:
            return {row[0] for row in conn.execute(f"select id from {table}")}
        except sqlite3.Error:
            return set()
        finally:
            conn.close()

    def read_legacy_master_codes(self) -> set[str]:
        if not self.legacy_path or not self.legacy_path.exists():
            return set()
        conn = sqlite3.connect(self.legacy_path)
        try:
            return {str(row[0]).strip() for row in conn.execute("select kode from master_akun")}
        except sqlite3.Error:
            return set()
        finally:
            conn.close()

    def read_legacy_count(self, table: str) -> int | None:
        if not self.legacy_path or not self.legacy_path.exists():
            return None
        conn = sqlite3.connect(self.legacy_path)
        try:
            return conn.execute(f"select count(*) from {table}").fetchone()[0]
        except sqlite3.Error:
            return None
        finally:
            conn.close()

    def count_orphans(self, table: str, fk_field: str, ref_table: str) -> int:
        with connection.cursor() as cursor:
            cursor.execute(
                f"select count(*) from {table} left join {ref_table} on {table}.{fk_field} = {ref_table}.id "
                f"where {table}.{fk_field} is not null and {ref_table}.id is null"
            )
            return cursor.fetchone()[0]

    def extract_cell(self, message: str) -> str:
        match = re.search(r"Cell ([A-Z]+[0-9]+)", message)
        return match.group(1) if match else ""

    def column_hint(self, cell: str) -> str:
        if not cell:
            return "-"
        letters = re.match(r"([A-Z]+)", cell).group(1)
        return {
            "F": "Tanggal SPM pada sheet D_K",
            "E": "Nomor SPM / tanggal tergantung sheet",
            "G": "Jenis SPM / field sekitar tanggal",
        }.get(letters, f"Kolom {letters}")
