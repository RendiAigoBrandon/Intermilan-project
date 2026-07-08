from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from apps.core.import_utils import ImportStats, clean_text, dict_from_headers, parse_date, parse_decimal, parse_month, pick
from apps.dk.models import MasterAkun, TransactionDetail
from apps.documents.models import DocumentDriveLink

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


class Command(BaseCommand):
    help = "Import seed Excel INTERMILAN/KK_13xx secara aman ke database Django."

    def add_arguments(self, parser):
        parser.add_argument("--path", required=True, help="Path folder Database awal atau file .xlsx.")
        parser.add_argument("--commit", action="store_true", help="Jalankan import asli. Default adalah dry-run.")
        parser.add_argument("--skip-duplicates", action="store_true", default=True, help="Skip data duplikat. Default aktif.")
        parser.add_argument("--replace-confirmed", action="store_true", help="Update data existing jika duplikat ditemukan.")
        parser.add_argument("--limit-files", type=int, default=0, help="Batasi jumlah file Excel untuk audit cepat.")
        parser.add_argument("--limit-rows", type=int, default=0, help="Batasi jumlah baris per sheet.")

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError("openpyxl belum tersedia.")
        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"Path Excel tidak ditemukan: {path}")
        if options["replace_confirmed"] and not options["commit"]:
            raise CommandError("--replace-confirmed hanya boleh dipakai bersama --commit.")

        self.commit = options["commit"]
        self.replace = options["replace_confirmed"]
        self.limit_rows = options["limit_rows"]
        self.stats: list[ImportStats] = []

        files = [path] if path.is_file() else sorted(path.glob("*.xlsx"))
        files = [f for f in files if not f.name.startswith("~$")]
        if options["limit_files"]:
            files = files[: options["limit_files"]]

        self.stdout.write(self.style.WARNING("Mode: IMPORT ASLI") if self.commit else self.style.WARNING("Mode: DRY-RUN"))
        self.stdout.write(f"Excel files: {len(files)}")

        for file_path in files:
            if file_path.name.upper().startswith("KK_"):
                self.import_kk_file(file_path)
            elif file_path.name.upper() == "INTERMILAN.XLSX":
                self.audit_intermilan_file(file_path)

        self.print_summary()

    def import_kk_file(self, path: Path):
        satker_code = self.satker_from_filename(path)
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if "D_K" in wb.sheetnames:
                self.import_dk_sheet(wb["D_K"], path.name, satker_code)
            if "Upload" in wb.sheetnames:
                self.import_upload_sheet(wb["Upload"], path.name, satker_code)
        finally:
            wb.close()

    def audit_intermilan_file(self, path: Path):
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            stats = ImportStats(f"{path.name}:reference")
            for sheet_name in ("Dashboard", "Monitoring_Combine", "Data_Integrasi", "DB"):
                if sheet_name in wb.sheetnames:
                    stats.read += 1
            stats.skipped = stats.read
            self.stats.append(stats)
        finally:
            wb.close()

    def import_dk_sheet(self, ws, filename: str, satker_code: str):
        stats = ImportStats(f"{filename}:D_K")
        headers = None
        for row in ws.iter_rows(values_only=True):
            values = tuple(row)
            if not any(v not in (None, "") for v in values):
                continue
            if headers is None:
                normalized = [clean_text(v).lower() for v in values]
                if any("nomor spm" in h.lower() for h in normalized) and any("akun" in h.lower() for h in normalized):
                    headers = list(values)
                continue
            if self.limit_rows and stats.read >= self.limit_rows:
                break
            stats.read += 1
            row_data = dict_from_headers(headers, values)
            nomor_spm = clean_text(pick(row_data, "nomor spm"))
            akun = clean_text(pick(row_data, "akun"))
            if not nomor_spm and not akun:
                stats.skipped += 1
                continue
            no_kuitansi = clean_text(pick(row_data, "no kuitansi", "no kuitansi hanya untuk dana up ptup no spm"))
            no_drpp = clean_text(pick(row_data, "no drpp"))
            nilai_netto = parse_decimal(pick(row_data, "nilai netto", default=pick(row_data, "nilai bruto", "nilai")))
            duplicate = TransactionDetail.objects.filter(
                satker_code=satker_code,
                nomor_spm=nomor_spm,
                akun=akun,
                no_kuitansi=no_kuitansi,
                no_drpp=no_drpp,
                nilai_netto=nilai_netto,
            ).exists()
            if duplicate and not self.replace:
                stats.duplicates += 1
                stats.skipped += 1
                self.ensure_master_akun(akun, row_data, stats)
                continue
            if not self.commit:
                stats.success += 1
                self.ensure_master_akun(akun, row_data, stats)
                continue

            obj = TransactionDetail.objects.filter(
                satker_code=satker_code,
                nomor_spm=nomor_spm,
                akun=akun,
                no_kuitansi=no_kuitansi,
                no_drpp=no_drpp,
                nilai_netto=nilai_netto,
            ).first() if self.replace else None
            obj = obj or TransactionDetail()
            obj.satker_code = satker_code
            obj.akun = akun or "-"
            obj.kategori = self.resolve_kategori(row_data)
            obj.bulan_sp2d = parse_month(pick(row_data, "sp2d bulan", "bulan sp2d"))
            obj.cara_pembayaran = clean_text(pick(row_data, "cara pembayaran"))
            obj.nomor_spm = nomor_spm
            obj.tanggal_spm = parse_date(pick(row_data, "tanggal spm"))
            obj.jenis_spm = clean_text(pick(row_data, "jenis spm"))
            obj.no_kuitansi = no_kuitansi
            obj.no_drpp = no_drpp
            obj.deskripsi = clean_text(pick(row_data, "deskripsi", "uraian belanja per transaksi"))
            obj.nilai_bruto = parse_decimal(pick(row_data, "nilai bruto", "nilai"))
            obj.nilai_netto = nilai_netto
            obj.pembebanan = clean_text(pick(row_data, "pembebanan"))
            obj.fp = clean_text(pick(row_data, "fp"))
            obj.pph21 = parse_decimal(pick(row_data, "pph21"))
            obj.save()
            stats.success += 1
            self.ensure_master_akun(akun, row_data, stats)
        self.stats.append(stats)

    def import_upload_sheet(self, ws, filename: str, satker_code: str):
        stats = ImportStats(f"{filename}:Upload")
        headers = None
        for row in ws.iter_rows(values_only=True):
            values = tuple(row)
            if not any(v not in (None, "") for v in values):
                continue
            if headers is None:
                if any(clean_text(v).lower() in {"url", "link", "link google drive"} for v in values):
                    headers = list(values)
                continue
            if self.limit_rows and stats.read >= self.limit_rows:
                break
            stats.read += 1
            row_data = dict_from_headers(headers, values)
            key = clean_text(pick(row_data, "no spm kuitansi", "no spm", "kuitansi"))
            url = clean_text(pick(row_data, "url", "link", "link google drive"))
            if not key or not url:
                stats.skipped += 1
                continue
            transaction = (
                TransactionDetail.objects.filter(satker_code=satker_code, nomor_spm=key).first()
                or TransactionDetail.objects.filter(satker_code=satker_code, no_kuitansi=key).first()
                or TransactionDetail.objects.filter(satker_code=satker_code, no_drpp=key).first()
            )
            duplicate = DocumentDriveLink.objects.filter(satker_code=satker_code, google_drive_url=url).exists()
            if duplicate and not self.replace:
                stats.duplicates += 1
                stats.skipped += 1
                continue
            if not self.commit:
                stats.success += 1
                continue
            obj = DocumentDriveLink.objects.filter(satker_code=satker_code, google_drive_url=url).first() if self.replace else None
            obj = obj or DocumentDriveLink()
            obj.transaction_detail = transaction
            obj.satker_code = satker_code
            obj.nomor_spm = transaction.nomor_spm if transaction else key
            obj.no_kuitansi = transaction.no_kuitansi if transaction else ""
            obj.no_drpp = transaction.no_drpp if transaction else ""
            obj.jenis_dokumen = "Google Drive"
            obj.nama_file = key
            obj.google_drive_url = url
            obj.status = DocumentDriveLink.Status.AKTIF
            obj.catatan = f"Imported from {filename} sheet Upload"
            obj.save()
            stats.success += 1
        self.stats.append(stats)

    def ensure_master_akun(self, akun: str, row_data: dict, stats: ImportStats):
        akun = clean_text(akun)
        if not akun:
            return
        kategori = self.resolve_kategori(row_data)
        if not self.commit:
            return
        MasterAkun.objects.get_or_create(
            kode=akun,
            defaults={
                "nama_akun": clean_text(pick(row_data, "nama akun")) or f"Akun {akun}",
                "kategori": kategori,
                "source": "excel_seed",
            },
        )

    def resolve_kategori(self, row_data: dict) -> str:
        kategori = clean_text(pick(row_data, "kategori", "kategori otomatis"))
        if kategori:
            return kategori
        akun = clean_text(pick(row_data, "akun"))
        if akun.startswith("51"):
            return "Belanja Pegawai"
        if akun.startswith("52"):
            return "Belanja Barang/Jasa"
        if akun.startswith("53"):
            return "Belanja Modal"
        return ""

    def satker_from_filename(self, path: Path) -> str:
        stem = path.stem.upper()
        if stem.startswith("KK_"):
            return stem.replace("KK_", "").strip()
        return ""

    def print_summary(self):
        self.stdout.write("")
        self.stdout.write(self.style.HTTP_INFO("Ringkasan import Excel seed"))
        for stats in self.stats:
            self.stdout.write(
                f"- {stats.source}: read={stats.read}, success={stats.success}, "
                f"skip={stats.skipped}, duplicate={stats.duplicates}, failed={stats.failed}"
            )
