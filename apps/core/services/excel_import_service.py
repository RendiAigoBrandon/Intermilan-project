"""Multi-Satker Excel Import Service - Updated for unit_code/satker_code distinction."""
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from typing import Optional

from django.db import transaction as db_transaction

from apps.core.models import TransactionPackage, TransactionProvenance, SatkerMaster
from apps.core.services.transaction_services import (
    find_or_create_package,
    normalize_nomor_spm,
    normalize_tahun,
)
from apps.core.import_utils import (
    parse_decimal,
    parse_date,
    parse_month,
    clean_text,
    pick,
    dict_from_headers,
    normalize_header,
)


@dataclass
class KKExcelRow:
    satker_code: str = ""  # Official 6-digit
    unit_code: str = ""  # 4-digit from filename
    tahun: int = 0
    bulan_sp2d: int = 0
    akun: str = ""
    kategori: str = ""
    cara_pembayaran: str = ""
    nomor_spm: str = ""
    tanggal_spm: date = None
    jenis_spm: str = ""
    no_kuitansi: str = ""
    no_drpp: str = ""
    deskripsi: str = ""
    nilai_bruto: Decimal = Decimal("0")
    nilai_netto: Decimal = Decimal("0")
    pembebanan: str = ""
    fp: str = ""
    pph21: Decimal = Decimal("0")


@dataclass
class KKExcelImportResult:
    source_file: str
    unit_code: str = ""
    satker_code: str = ""
    read: int = 0
    new_rows: int = 0
    updated_rows: int = 0
    unchanged_rows: int = 0
    rejected_rows: int = 0
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)

    def add_error(self, message: str):
        self.rejected_rows += 1
        self.errors.append(message)

    def add_warning(self, message: str):
        self.warnings.append(message)

    @property
    def summary(self) -> str:
        return (
            f"File {self.source_file} (unit {self.unit_code} -> satker {self.satker_code}): "
            f"{self.new_rows} baru, {self.updated_rows} diperbarui, "
            f"{self.unchanged_rows} tidak berubah, {self.rejected_rows} ditolak"
        )


def extract_unit_code_from_filename(filename: str) -> Optional[str]:
    """Extract 4-digit unit code from filename like KK_1300.xlsx."""
    match = re.search(r"[Kk][Kk]_?(\d{4})", filename)
    return match.group(1) if match else None


def get_official_satker_code(unit_code: str) -> Optional[str]:
    """Get official 6-digit satker_code from 4-digit unit_code via SatkerMaster."""
    return SatkerMaster.get_satker_code(unit_code)


def parse_kk_excel_row(row_dict: dict, satker_code: str = None, unit_code: str = None) -> Optional[KKExcelRow]:
    """Parse a row from KK Excel into KKExcelRow using official satker_code."""
    try:
        tahun_raw = pick(row_dict, "tahun", "tahun anggaran", "ta", default="")
        tahun = normalize_tahun(tahun_raw)

        bulan_raw = pick(row_dict, "bulan", "bulan sp2d", "bulan_sp2d", default="")
        bulan_sp2d = parse_month(bulan_raw) or 0

        akun = clean_text(pick(row_dict, "akun", "kode akun", "kdakun", default=""))
        kategori = clean_text(pick(row_dict, "kategori", "kat", default=""))
        cara_pembayaran = clean_text(pick(row_dict, "cara pembayaran", "carabayar", "cara bayar", default=""))
        nomor_spm = normalize_nomor_spm(pick(row_dict, "nomor spm", "nomor_spm", "no spm", "no_spm", default=""))
        tanggal_spm = parse_date(pick(row_dict, "tanggal spm", "tanggal_spm", "tgl spm", default=None))
        jenis_spm = clean_text(pick(row_dict, "jenis spm", "jenis_spm", "jns spm", default=""))
        no_kuitansi = normalize_nomor_spm(pick(row_dict, "no kuitansi", "no_kuitansi", "kuitansi", "kw", default=""))
        no_drpp = normalize_nomor_spm(pick(row_dict, "no drpp", "no_drpp", "drpp", default=""))
        nilai_bruto = parse_decimal(pick(row_dict, "nilai bruto", "nilai_bruto", "bruto", "jumlah bruto", default=0))
        nilai_netto = parse_decimal(pick(row_dict, "nilai netto", "nilai_netto", "netto", "jumlah netto", default=0))
        if nilai_netto == 0 and nilai_bruto > 0:
            nilai_netto = nilai_bruto
        deskripsi = clean_text(pick(row_dict, "deskripsi", "rincian", "keterangan", default=""))
        pembebanan = clean_text(pick(row_dict, "pembebanan", "beban", default=""))
        fp = clean_text(pick(row_dict, "fp", "faktur", "nomor fp", default=""))
        pph21 = parse_decimal(pick(row_dict, "pph21", "pph 21", "ppn", default=0))

        if not satker_code and not unit_code:
            return None

        return KKExcelRow(
            satker_code=satker_code or "",
            unit_code=unit_code or "",
            tahun=tahun,
            bulan_sp2d=bulan_sp2d,
            akun=akun,
            kategori=kategori,
            cara_pembayaran=cara_pembayaran,
            nomor_spm=nomor_spm,
            tanggal_spm=tanggal_spm,
            jenis_spm=jenis_spm,
            no_kuitansi=no_kuitansi,
            no_drpp=no_drpp,
            deskripsi=deskripsi,
            nilai_bruto=nilai_bruto,
            nilai_netto=nilai_netto,
            pembebanan=pembebanan,
            fp=fp,
            pph21=pph21,
        )
    except Exception:
        return None


def get_dk_upsert_key(row: KKExcelRow) -> dict:
    """Get unique key using official 6-digit satker_code."""
    return {
        "satker_code": row.satker_code,
        "tahun": row.tahun,
        "nomor_spm": row.nomor_spm,
        "no_drpp": row.no_drpp,
        "no_kuitansi": row.no_kuitansi,
        "akun": row.akun,
    }


def find_existing_dk_row(row: KKExcelRow):
    """Find existing D_K row matching upsert key."""
    from apps.dk.models import TransactionDetail
    key = get_dk_upsert_key(row)
    if not all([key.get("satker_code"), key.get("tahun"), key.get("nomor_spm")]):
        return None
    return TransactionDetail.objects.filter(**key).first()


def update_dk_row_from_kk(row: KKExcelRow, existing_row=None, source_file: str = ""):
    """Update D_K row from KK Excel data."""
    from apps.dk.models import TransactionDetail

    created = False
    updated = False

    if existing_row:
        dk_row = existing_row
    else:
        dk_row = TransactionDetail()
        created = True

    dk_row.satker_code = row.satker_code
    dk_row.tahun = row.tahun
    dk_row.akun = row.akun
    dk_row.kategori = row.kategori
    dk_row.bulan_sp2d = row.bulan_sp2d
    dk_row.cara_pembayaran = row.cara_pembayaran
    dk_row.nomor_spm = row.nomor_spm
    dk_row.tanggal_spm = row.tanggal_spm
    dk_row.jenis_spm = row.jenis_spm
    dk_row.no_kuitansi = row.no_kuitansi
    dk_row.no_drpp = row.no_drpp
    dk_row.deskripsi = row.deskripsi
    dk_row.nilai_bruto = row.nilai_bruto
    dk_row.nilai_netto = row.nilai_netto
    dk_row.pembebanan = row.pembebanan
    dk_row.fp = row.fp
    dk_row.pph21 = row.pph21
    dk_row.source_type = TransactionProvenance.SourceType.KK_EXCEL
    dk_row.source_filename = source_file

    if existing_row:
        for field_name in ["nilai_bruto", "nilai_netto", "tanggal_spm", "jenis_spm"]:
            if getattr(existing_row, field_name) != getattr(row, field_name):
                updated = True
                break

    dk_row.save()

    if row.nomor_spm and row.tahun and row.satker_code:
        try:
            package, _ = find_or_create_package(
                satker_code=row.satker_code,
                tahun=row.tahun,
                nomor_spm=row.nomor_spm,
            )
            dk_row.transaction_package = package
            dk_row.save()
        except Exception:
            pass

    return created, updated


def import_kk_excel_file(filepath: str, source_file: str = "", user=None) -> KKExcelImportResult:
    """Import D_K from KK_XXXX.xlsx using SatkerMaster mapping."""
    import openpyxl

    result = KKExcelImportResult(
        source_file=source_file or filepath,
        unit_code=extract_unit_code_from_filename(source_file or filepath) or "",
    )

    if not result.unit_code:
        result.add_error("Tidak dapat mengekstrak kode unit dari nama file")
        return result

    result.satker_code = get_official_satker_code(result.unit_code) or ""
    if not result.satker_code:
        result.add_error(
            f"Kode unit {result.unit_code} tidak ditemukan di master satker. "
            f"Seed data SatkerMaster terlebih dahulu."
        )
        return result

    unit_code = result.unit_code
    official_satker_code = result.satker_code

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        headers = [normalize_header(cell.value) if cell.value else "" for cell in ws[1]]

        # Detect workbook satker column
        satker_col_idx = None
        for idx, header in enumerate(headers):
            if header in ("satker", "kode satker", "kdsatker", "kode_satker"):
                satker_col_idx = idx
                break

        # Validate workbook satker against official satker_code
        if satker_col_idx is not None:
            workbook_satker = None
            for row in ws.iter_rows(min_row=2, max_row=min(10, ws.max_row), values_only=True):
                if row[satker_col_idx]:
                    workbook_satker = clean_text(str(row[satker_col_idx]).strip()).zfill(6)
                    break

            if workbook_satker and workbook_satker != official_satker_code:
                wb.close()
                result.add_error(
                    f"File {source_file} seharusnya milik satker {official_satker_code}, "
                    f"tetapi isi workbook menunjukkan kode {workbook_satker}. "
                    f"Periksa file sebelum melanjutkan."
                )
                return result

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            result.read += 1
            row_dict = dict_from_headers(headers, row)
            parsed = parse_kk_excel_row(row_dict, satker_code=official_satker_code, unit_code=unit_code)

            if not parsed:
                result.add_error(f"Baris {row_num}: Tidak dapat memproses data")
                continue

            if not parsed.nomor_spm and not parsed.no_kuitansi and not parsed.no_drpp:
                result.add_error(f"Baris {row_num}: Tidak ada nomor SPM, DRPP, atau kuitansi")
                continue

            existing = find_existing_dk_row(parsed)
            if existing and not _has_significant_changes(existing, parsed):
                result.unchanged_rows += 1
                continue

            try:
                with db_transaction.atomic():
                    created, updated = update_dk_row_from_kk(parsed, existing, result.source_file)
                    if created:
                        result.new_rows += 1
                    elif updated:
                        result.updated_rows += 1
                    else:
                        result.unchanged_rows += 1
            except Exception as e:
                result.add_error(f"Baris {row_num}: {str(e)}")

        wb.close()
    except Exception as e:
        result.add_error(f"Gagal membaca file: {str(e)}")

    return result


def _has_significant_changes(existing, parsed: KKExcelRow) -> bool:
    if existing.nilai_bruto != parsed.nilai_bruto:
        return True
    if existing.nilai_netto != parsed.nilai_netto:
        return True
    if existing.tanggal_spm != parsed.tanggal_spm:
        return True
    if existing.jenis_spm != parsed.jenis_spm:
        return True
    return False


def import_multiple_kk_files(filepaths: list, user=None) -> list:
    results = []
    for filepath in filepaths:
        source_file = filepath.split("/")[-1].split("\\")[-1]
        result = import_kk_excel_file(filepath, source_file, user)
        results.append(result)
    return results
