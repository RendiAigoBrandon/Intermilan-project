import os
import sys
from openpyxl import load_workbook

data_dir = r'c:\Users\muall\Documents\INTERMILAN PROJECT\data_sources\Database awal'
kk_files = []
for root, dirs, files in os.walk(data_dir):
    for f in files:
        if f.startswith('KK_') and f.endswith('.xlsx') and not f.startswith('~$'):
            kk_files.append(os.path.join(root, f))

print(f"Found {len(kk_files)} files to process.")
sys.stdout.flush()

unique_accounts = {}
total_transactions = 0
files_read = 0

for file_path in kk_files:
    fname = os.path.basename(file_path)
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if 'D_K' not in wb.sheetnames:
            print(f"No D_K sheet in {fname}")
            sys.stdout.flush()
            wb.close()
            continue
            
        ws = wb['D_K']
        files_read += 1
        headers = None
        
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=1000)):
            if not any(v not in (None, '') for v in row):
                continue
            
            if headers is None:
                normalized = [str(v).lower().strip() if v else '' for v in row]
                if any('akun' in h for h in normalized):
                    headers = normalized
                continue
            
            row_data = dict(zip(headers, row))
            akun_raw = None
            for h in headers:
                if 'akun' in h:
                    akun_raw = row_data[h]
            
            if akun_raw is not None and str(akun_raw).strip() != '' and str(akun_raw).strip() != 'None':
                akun_clean = str(akun_raw).strip()
                if akun_clean.endswith('.0'):
                    akun_clean = akun_clean[:-2]
                
                # Check if it's actually an account code (numeric or starting with 5/8)
                if not (akun_clean.startswith('5') or akun_clean.startswith('8')):
                    continue
                    
                total_transactions += 1
                
                kategori = ''
                if akun_clean.startswith('51'): kategori = 'Belanja Pegawai'
                elif akun_clean.startswith('52'): kategori = 'Belanja Barang'
                elif akun_clean.startswith('53'): kategori = 'Belanja Modal'
                elif akun_clean.startswith('825'): kategori = 'Transaksi UP'
                    
                if akun_clean not in unique_accounts:
                    unique_accounts[akun_clean] = {
                        'kode': akun_clean,
                        'nama': '',
                        'kategori': kategori,
                        'sumber': set([fname]),
                    }
                else:
                    unique_accounts[akun_clean]['sumber'].add(fname)
        wb.close()
        print(f"Processed {fname}")
        sys.stdout.flush()
    except Exception as e:
        print(f"Error reading {fname}: {e}")
        sys.stdout.flush()

report = []
report.append('# Rekap Akun Keuangan Database Awal\n')
report.append(f'- Jumlah file KK yang dibaca: {files_read}')
report.append(f'- Jumlah transaksi yang dianalisis: {total_transactions}')
report.append(f'- Jumlah akun unik: {len(unique_accounts)}\n')
report.append('| No | Kode Akun | Nama Akun | Kategori | Sumber File KK |')
report.append('|---|---|---|---|---|')

for i, (kode, data) in enumerate(sorted(unique_accounts.items(), key=lambda x: x[0]), 1):
    sumber_list = ', '.join(sorted(list(data['sumber']))[:3])
    if len(data['sumber']) > 3:
        sumber_list += '...'
    report.append(f"| {i} | {kode} | {data['nama']} | {data['kategori']} | {sumber_list} |")

report.append('\n## Kesimpulan\n')
report.append('**Jika database awal ini dijadikan sumber MasterAkun INTERMILAN, daftar akun apa saja yang harus dimasukkan?**\n')
report.append('Daftar akun yang WAJIB dimasukkan adalah seluruh akun unik di atas. Namun, ada beberapa hal krusial:\n')
report.append('1. **Kode Invalid**: Kode seperti `51XXXX` dan `51XXX` harus dibuang atau diubah karena tidak spesifik (ini adalah kode kelompok).\n')
report.append('2. **Nama Akun**: Nama akun tidak tercantum secara spesifik di sheet transaksi (D_K). Oleh karena itu, kita harus memberikan nama default atau menarik nama tersebut dari referensi dokumen lain (misalnya sheet nama kategori).\n')
report.append('3. **Format**: Format seperti `.0` telah dinormalisasi.\n')
report.append('\nDatabase awal ini cukup memberikan representasi akun aktif yang sedang digunakan satker, namun tetap butuh pelengkap nama akun agar tidak kosong di aplikasi.')

artifact_path = r'C:\Users\muall\.gemini\antigravity-ide\brain\d8cfb0d7-ce45-4337-8751-d70f5d18b9d2\DATABASE_AWAL_EXCEL_AUDIT.md'
with open(artifact_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(report))

print('DONE')
sys.stdout.flush()
